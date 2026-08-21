"""
processos/relatorios.py

Geração dos entregáveis do processo a partir do dicionário consolidado
pelo ai_processor.merge_results():

    * Mapa Comparativo de Preços (.xlsx), sobre o modelo institucional.
    * Relatório do processo (.odt), com a trilha de auditoria da extração.

Nenhuma função aqui acessa `request` nem grava no banco: recebem o processo
e os dados, devolvem o caminho do arquivo gerado.
"""

import logging
import os

import openpyxl
from django.conf import settings
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .services import chave_empresa, para_float

logger = logging.getLogger(__name__)

# Layout do modelo institucional Mapa_Comparativo_Base.xlsx.
LINHA_CABECALHO = 3
LINHA_PRIMEIRO_ITEM = 4
COLUNAS_FIXAS = ['ITEM', 'PI', 'NOME EM PORTUGUÊS', 'QTDE', 'UF', 'PAINEL DE PREÇO']
COLUNA_PRIMEIRA_EMPRESA = len(COLUNAS_FIXAS) + 1        # G
MAXIMO_DE_EMPRESAS = 20                                 # G..Z
COLUNA_VALOR_UNITARIO = 27                              # AA
COLUNA_VALOR_TOTAL = 28                                 # AB

FORMATO_MOEDA = 'R$ #,##0.00'
PASTA_GERADOS = ('processos', 'gerados')


# ==================== RESUMO DA EXTRAÇÃO ====================

def resumo_extracao(resultados):
    """Consolida a rota de extração de cada arquivo e o custo da IA.

    A rota ('modelo/deterministico', 'pdf/pdf-text+ocr-local', 'imagem'...)
    é o que permite auditar de onde veio cada número do mapa. Sem isso a
    informação morria dentro do merge_results.
    """
    arquivos = []
    tokens_entrada = 0
    tokens_saida = 0
    custo = 0.0

    for registro in resultados:
        resultado = registro.get('ai_result') or {}
        uso = resultado.get('uso') or {}
        tokens_entrada += uso.get('entrada') or 0
        tokens_saida += uso.get('saida') or 0
        custo += para_float(uso.get('custo_usd')) or 0.0

        arquivos.append({
            'arquivo': registro.get('filename', ''),
            'rota': resultado.get('rota') or '',
            'status': resultado.get('status') or '',
            'erro': resultado.get('error') or '',
            'truncado': bool(resultado.get('truncado')),
        })

    def nomes(condicao):
        return sorted({a['arquivo'] for a in arquivos if condicao(a)})

    return {
        'arquivos': arquivos,
        'tokens_entrada': tokens_entrada,
        'tokens_saida': tokens_saida,
        'custo_usd': round(custo, 6),
        'ocr_local': getattr(settings, 'OCR_LOCAL', True),
        'com_ocr_local': nomes(lambda a: 'ocr-local' in a['rota']),
        'sem_ia': nomes(lambda a: 'deterministico' in a['rota']),
        'com_falha': nomes(lambda a: a['status'] not in ('success', '')),
        'truncados': nomes(lambda a: a['truncado']),
    }


def avisos_da_extracao(extracao):
    """Traduz o resumo técnico em avisos para quem vai conferir o mapa."""
    avisos = []

    if extracao['com_ocr_local']:
        avisos.append(
            'OCR local (Tesseract) aplicado em '
            + ', '.join(extracao['com_ocr_local'])
            + ': confira preços e PI desses arquivos dígito a dígito — '
              'erro de OCR em número não é detectável pelo sistema.'
        )

    for nome in extracao['com_falha']:
        detalhe = next((a['erro'] for a in extracao['arquivos']
                        if a['arquivo'] == nome and a['erro']),
                       'motivo não informado')
        avisos.append(f'{nome}: não foi possível extrair ({detalhe}).')

    for nome in extracao['truncados']:
        avisos.append(f'{nome}: resposta truncada por limite de tokens — '
                      f'itens podem ter ficado de fora.')

    return avisos


# ==================== MAPA COMPARATIVO (XLSX) ====================

def _caminho_de_saida(processo, extensao):
    pasta = os.path.join(settings.MEDIA_ROOT, *PASTA_GERADOS)
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, f'mapa_comparativo_{processo.numero_slug}.{extensao}')


def _abrir_modelo():
    """Abre o modelo institucional ou cria uma planilha vazia equivalente."""
    caminho = os.path.join(settings.BASE_DIR, 'static-assets', 'modelos',
                           'Mapa_Comparativo_Base.xlsx')
    if os.path.exists(caminho):
        return openpyxl.load_workbook(caminho)

    logger.warning('Modelo %s não encontrado; gerando planilha sem formatação base.',
                   caminho)
    planilha = openpyxl.Workbook()
    planilha.active.title = 'MAPA COMPARATIVO DO PROCESSO'
    return planilha


def _nome_da_empresa(empresa, posicao):
    if isinstance(empresa, dict):
        return (empresa.get('nome') or '').strip() or f'EMPRESA{posicao}'
    return str(empresa) or f'EMPRESA{posicao}'


def preencher_mapa_comparativo(processo, dados_ai):
    """Preenche o Mapa Comparativo e devolve o caminho do arquivo gerado."""
    planilha = _abrir_modelo()
    aba = planilha.active

    empresas = (dados_ai.get('empresas') or [])[:MAXIMO_DE_EMPRESAS]
    excedentes = len(dados_ai.get('empresas') or []) - len(empresas)
    if excedentes > 0:
        aviso = (f'{excedentes} fornecedor(es) não couberam no mapa: o modelo '
                 f'comporta {MAXIMO_DE_EMPRESAS} colunas de empresa.')
        dados_ai.setdefault('avisos_gerais', []).append(aviso)
        logger.warning('Processo %s: %s', processo.numero, aviso)

    estilos = _estilos()
    _escrever_cabecalho(aba, empresas, estilos)
    _escrever_itens(aba, dados_ai.get('itens') or [], empresas, estilos)
    _ajustar_larguras(aba)

    caminho = _caminho_de_saida(processo, 'xlsx')
    planilha.save(caminho)
    return caminho


def _estilos():
    return {
        'preenchimento': PatternFill(start_color='4472C4', end_color='4472C4',
                                     fill_type='solid'),
        'fonte': Font(bold=True, color='FFFFFF', size=10),
        'borda': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'centro': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }


def _escrever_cabecalho(aba, empresas, estilos):
    titulos = list(COLUNAS_FIXAS)
    titulos += [_nome_da_empresa(e, i)[:30] for i, e in enumerate(empresas, start=1)]

    # As colunas de valor ficam em posição fixa no modelo (AA e AB).
    for coluna, titulo in enumerate(titulos, start=1):
        celula = aba.cell(row=LINHA_CABECALHO, column=coluna, value=titulo)
        celula.font = estilos['fonte']
        celula.fill = estilos['preenchimento']
        celula.alignment = estilos['centro']
        celula.border = estilos['borda']

    for coluna, titulo in ((COLUNA_VALOR_UNITARIO, 'VALOR UNITARIO ESTIMADO'),
                           (COLUNA_VALOR_TOTAL, 'VALOR TOTAL')):
        celula = aba.cell(row=LINHA_CABECALHO, column=coluna, value=titulo)
        celula.font = estilos['fonte']
        celula.fill = estilos['preenchimento']
        celula.alignment = estilos['centro']
        celula.border = estilos['borda']


def _escrever_itens(aba, itens, empresas, estilos):
    for merged in list(aba.merged_cells.ranges):
        if merged.min_row >= LINHA_PRIMEIRO_ITEM:
            aba.unmerge_cells(str(merged))
    for posicao, item in enumerate(itens, start=1):
        linha = LINHA_PRIMEIRO_ITEM + posicao - 1

        aba.cell(row=linha, column=1, value=item.get('item', posicao))
        aba.cell(row=linha, column=2, value=item.get('pi', ''))
        aba.cell(row=linha, column=3, value=item.get('nome_em_portugues', ''))
        aba.cell(row=linha, column=4, value=para_float(item.get('qtde')))
        aba.cell(row=linha, column=5, value=item.get('uf', ''))
        _escrever_valor(aba, linha, 6, item.get('painel_preco'))

        _escrever_precos(aba, linha, item, empresas, estilos)

        _escrever_valor(aba, linha, COLUNA_VALOR_UNITARIO,
                        item.get('valor_unitario_estimado'))
        _escrever_valor(aba, linha, COLUNA_VALOR_TOTAL, item.get('valor_total'))

        for coluna in range(1, COLUNA_VALOR_TOTAL + 1):
            aba.cell(row=linha, column=coluna).border = estilos['borda']


def _escrever_precos(aba, linha, item, empresas, estilos):
    """Escreve a cotação de cada empresa na coluna correspondente.

    A IA devolve {NOME DA EMPRESA: preço}; o JSON antigo devolvia
    {empresaN: preço}. As duas formas são aceitas.
    """
    cotacoes = item.get('empresas') or {}
    por_nome = {chave_empresa(k): v for k, v in cotacoes.items()}

    for posicao, empresa in enumerate(empresas, start=1):
        coluna = COLUNA_PRIMEIRA_EMPRESA + posicao - 1
        nome = _nome_da_empresa(empresa, posicao)

        valor = por_nome.get(chave_empresa(nome))
        if valor in (None, ''):
            valor = cotacoes.get(f'empresa{posicao}')

        celula = _escrever_valor(aba, linha, coluna, valor)
        celula.alignment = estilos['centro']


def _escrever_valor(aba, linha, coluna, valor):
    """Grava como número quando for número, para a planilha poder calcular.

    Gravar preço como texto é o erro que faz o somatório do mapa voltar zero
    e obriga a conferência manual de tudo.
    """
    celula = aba.cell(row=linha, column=coluna)
    numero = para_float(valor)

    if numero is not None:
        celula.value = numero
        celula.number_format = FORMATO_MOEDA
    else:
        celula.value = '' if valor in (None, '') else str(valor)
    return celula


def _ajustar_larguras(aba):
    for coluna in range(1, COLUNA_VALOR_TOTAL + 1):
        letra = get_column_letter(coluna)
        estreita = COLUNA_PRIMEIRA_EMPRESA <= coluna < COLUNA_VALOR_UNITARIO
        aba.column_dimensions[letra].width = 12 if estreita else 20


# ==================== RELATÓRIO (ODT) ====================

def gerar_planilha_odt(processo, dados_ai):
    """Gera o relatório .odt do processo e devolve o caminho do arquivo."""
    from odf.opendocument import OpenDocumentText
    from odf.text import H, P

    documento = OpenDocumentText()

    def titulo(texto, nivel=2):
        documento.text.addElement(H(outlinelevel=nivel, text=texto))

    def paragrafo(texto=''):
        documento.text.addElement(P(text=texto))

    titulo(f'Mapa Comparativo - Processo {processo.numero}', nivel=1)

    titulo('Informações do Processo')
    paragrafo(f'Número: {processo.numero}')
    paragrafo(f'Descrição: {processo.descricao}')
    paragrafo(f'Valor Estimado: R$ {processo.valor_estimado:.2f}')
    paragrafo(f'Data: {processo.data_abertura.strftime("%d/%m/%Y")}')
    paragrafo()

    _secao_empresas(dados_ai, titulo, paragrafo)
    _secao_itens(dados_ai, titulo, paragrafo)
    _secao_extracao(dados_ai, titulo, paragrafo)
    _secao_perguntas_e_avisos(dados_ai, titulo, paragrafo)

    caminho = _caminho_de_saida(processo, 'odt')
    documento.save(caminho)
    return caminho


def _secao_empresas(dados_ai, titulo, paragrafo):
    empresas = dados_ai.get('empresas') or []
    if not empresas:
        return

    titulo('Empresas Participantes')
    for empresa in empresas:
        paragrafo(f'- {empresa.get("nome", "N/A")}')
        if empresa.get('cnpj'):
            paragrafo(f'  CNPJ: {empresa["cnpj"]}')

        tipo = empresa.get('tipo_resposta') or 'cotacao'
        if tipo != 'cotacao':
            rotulo = 'Declinou' if tipo == 'declinio' else 'Apenas dúvida/esclarecimento'
            motivo = empresa.get('motivo_declinio') or ''
            paragrafo(f'  {rotulo}{": " + motivo if motivo else ""}')

        if empresa.get('valor_global'):
            paragrafo(f'  Valor: {empresa["valor_global"]}')
        paragrafo()


def _secao_itens(dados_ai, titulo, paragrafo):
    itens = dados_ai.get('itens') or []
    if not itens:
        return

    titulo('Itens do Processo')
    for item in itens:
        paragrafo(f'Item {item.get("item", "")}: {item.get("nome_em_portugues", "")}')
        paragrafo(f'  PI: {item.get("pi", "")}')
        paragrafo(f'  Quantidade: {item.get("qtde", "")} {item.get("uf", "")}')

        if item.get('valor_unitario_estimado'):
            paragrafo(f'  Valor Unitário: {item["valor_unitario_estimado"]}')
        if item.get('valor_total'):
            paragrafo(f'  Valor Total: {item["valor_total"]}')

        cotacoes = {nome: valor for nome, valor in (item.get('empresas') or {}).items()
                    if valor not in (None, '')}
        if cotacoes:
            paragrafo('  Cotações:')
            for nome, valor in cotacoes.items():
                paragrafo(f'    {nome}: {valor}')
        else:
            paragrafo('  Sem cotação')
        paragrafo()


def _secao_extracao(dados_ai, titulo, paragrafo):
    extracao = dados_ai.get('extracao') or {}
    if not extracao.get('arquivos'):
        return

    titulo('Origem da Extração')
    for arquivo in extracao['arquivos']:
        linha = f'- {arquivo.get("arquivo", "")}: {arquivo.get("rota") or "n/d"}'
        if arquivo.get('status') and arquivo['status'] != 'success':
            linha += f' [{arquivo["status"]}: {arquivo.get("erro", "")}]'
        if arquivo.get('truncado'):
            linha += ' [resposta truncada por limite de tokens]'
        paragrafo(linha)

    if extracao.get('com_ocr_local'):
        paragrafo()
        paragrafo('ATENÇÃO: OCR local (Tesseract) foi aplicado em '
                  + ', '.join(extracao['com_ocr_local'])
                  + '. Preços e PI desses arquivos exigem conferência dígito a '
                    'dígito antes da homologação da pesquisa.')

    paragrafo(f'Tokens: {extracao.get("tokens_entrada", 0)} entrada / '
              f'{extracao.get("tokens_saida", 0)} saída. '
              f'Custo estimado: USD {extracao.get("custo_usd", 0)}.')
    paragrafo()


def _secao_perguntas_e_avisos(dados_ai, titulo, paragrafo):
    perguntas = dados_ai.get('perguntas') or []
    if perguntas:
        titulo('Perguntas dos Fornecedores')
        for pergunta in perguntas:
            paragrafo(f'- {pergunta.get("empresa", "")}: {pergunta.get("pergunta", "")}')
        paragrafo()

    avisos = dados_ai.get('avisos_gerais') or []
    if avisos:
        titulo('Avisos do Processamento')
        for aviso in avisos:
            paragrafo(f'- {aviso}')
