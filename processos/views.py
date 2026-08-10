from django.shortcuts import render, redirect
from django.utils import timezone
from .models import Processo
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse, FileResponse
from django.conf import settings
import re
import os
import zipfile
import tempfile
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
# >>> ALTERADO: parse_modelo_proposta é função de módulo, não método da classe
from .ai_processor import AIProcessor, parse_modelo_proposta
from django.db import IntegrityError
# >>> ALTERADO: linha_base (2ª fase) e chave_empresa (casamento das cotações)
from .services import montar_resumo, linha_base, chave_empresa
from .persistencia import salvar_dados_ai
from django.contrib.auth.decorators import login_required

# ==================== VIEWS PRINCIPAIS ====================

@login_required
def processos(request):
    processos = Processo.objects.filter(usuario=request.user)
    ordem = request.GET.get("ordem")
    data = request.GET.get("data")
    busca = request.GET.get("busca")

    if busca:
        palavras = busca.split()
        for palavra in palavras:
            processos = processos.filter(Q(numero__icontains=palavra) | Q(descricao__icontains=palavra))

    hoje = timezone.now().date()
    if data == "hoje":
        processos = processos.filter(data_abertura=hoje)
    elif data == "mes":
        processos = processos.filter(data_abertura__month=hoje.month, data_abertura__year=hoje.year)
    elif data == "ano":
        processos = processos.filter(data_abertura__year=hoje.year)

    if ordem == "antigos":
        processos = processos.order_by("id")
    elif ordem == "numero":
        processos = processos.order_by("numero")
    elif ordem == "recentes":
        processos = processos.order_by("-id")
    elif ordem == "maior_valor":
        processos = processos.order_by("-valor_estimado")
    elif ordem == "menor_valor":
        processos = processos.order_by("valor_estimado")

    paginator = Paginator(processos, 10)
    page = request.GET.get("page")
    processos = paginator.get_page(page)

    return render(
        request,
        "processos.html",
        {
            "processos": processos,
            "busca": busca,
            "ordem": ordem,
            "data": data,
        }
    )

# ==================== FUNÇÕES DE PROCESSAMENTO ====================

def contar_emails(diretorio):
    """Conta os .eml/.msg extraídos do pacote (respostas dos fornecedores)."""
    total = 0
    for raiz, _, arquivos in os.walk(diretorio):
        total += sum(1 for nome in arquivos if nome.lower().endswith(('.eml', '.msg')))
    return total


# >>> NOVO: consolida a "rota" de cada arquivo (a que o ai_processor devolve em
# resultado['rota']: 'modelo/deterministico', 'pdf/pdf-text', 'pdf/pdf-text+ocr-local',
# 'pdf/native', 'imagem', 'texto/eml'...). Sem isso essa informação morria no merge.
def resumo_extracao(results):
    arquivos = []
    tokens_entrada = tokens_saida = 0
    custo = 0.0

    for registro in results:
        resultado = registro.get('ai_result') or {}
        uso = resultado.get('uso') or {}
        tokens_entrada += uso.get('entrada') or 0
        tokens_saida += uso.get('saida') or 0
        try:
            custo += float(uso.get('custo_usd') or 0)
        except (TypeError, ValueError):
            pass
        arquivos.append({
            'arquivo': registro.get('filename', ''),
            'rota': resultado.get('rota', '') or '',
            'status': resultado.get('status', '') or '',
            'erro': resultado.get('error', '') or '',
            'truncado': bool(resultado.get('truncado')),
        })

    def nomes(condicao):
        return sorted({a['arquivo'] for a in arquivos if condicao(a)})

    return {
        'arquivos': arquivos,
        'tokens_entrada': tokens_entrada,
        'tokens_saida': tokens_saida,
        'custo_usd': round(custo, 6),
        'ocr_local': getattr(settings, 'OCR_LOCAL', True),
        'com_ocr_local': nomes(lambda a: 'ocr-local' in a['rota']),
        'sem_ia': nomes(lambda a: 'deterministico' in a['rota']),
        'com_falha': nomes(lambda a: a['status'] not in ('success', '')),
        'truncados': nomes(lambda a: a['truncado']),
    }


def preencher_mapa_comparativo(processo, dados_ai):
    """Preenche o modelo Mapa_Comparativo_Base.xlsx com os dados extraídos pela IA"""
    try:
        template_path = os.path.join(settings.BASE_DIR, 'static-assets', 'modelos', 'Mapa_Comparativo_Base.xlsx')

        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MAPA COMPARATIVO DO PROCESSO"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Cabeçalhos (linha 3 do template)
        headers = ['ITEM', 'PI', 'NOME EM PORTUGUÊS', 'QTDE', 'UF', 'PAINEL DE PREÇO']
        empresas = dados_ai.get('empresas', [])
        # >>> ALTERADO: o cabeçalho traz a razão social, não "EMPRESA1"
        for idx, empresa in enumerate(empresas[:20], 1):
            nome = (empresa.get('nome') or '').strip() if isinstance(empresa, dict) else str(empresa)
            headers.append((nome or f'EMPRESA{idx}')[:30])
        headers.extend(['VALOR UNITARIO ESTIMADO', 'VALOR TOTAL'])

        # Escreve cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Preenche dados
        itens = dados_ai.get('itens', [])
        current_row = 4

        for item_idx, item in enumerate(itens, 1):
            row = current_row + item_idx - 1

            # Colunas A-F: dados básicos
            ws.cell(row=row, column=1, value=item.get('item', item_idx))
            ws.cell(row=row, column=2, value=item.get('pi', ''))
            ws.cell(row=row, column=3, value=item.get('nome_em_portugues', ''))
            ws.cell(row=row, column=4, value=item.get('qtde', ''))
            ws.cell(row=row, column=5, value=item.get('uf', ''))
            ws.cell(row=row, column=6, value=item.get('painel_preco', ''))

            # Colunas G-Z: empresas
            # >>> ALTERADO: a IA devolve {NOME DA EMPRESA: preço}, não {empresaN: preço}
            empresas_data = item.get('empresas', {}) or {}
            por_nome = {chave_empresa(k): v for k, v in empresas_data.items()}
            for emp_idx, empresa in enumerate(empresas[:20], start=1):
                col = 6 + emp_idx
                nome = empresa.get('nome') if isinstance(empresa, dict) else empresa
                valor = por_nome.get(chave_empresa(nome))
                if valor in (None, ''):
                    valor = empresas_data.get(f'empresa{emp_idx}', '')   # JSON antigo
                ws.cell(row=row, column=col, value=valor)
                ws.cell(row=row, column=col).alignment = center

            # Colunas AA-AB: valores
            ws.cell(row=row, column=27, value=item.get('valor_unitario_estimado', ''))
            ws.cell(row=row, column=28, value=item.get('valor_total', ''))

            # Aplica bordas
            for col in range(1, 29):
                ws.cell(row=row, column=col).border = border

        # Ajusta largura das colunas
        for col in range(1, 29):
            col_letter = get_column_letter(col)
            if col <= 6 or col >= 27:
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 12

        # Salva
        filename = f"mapa_comparativo_{processo.numero_slug}.xlsx"   # >>> ALTERADO: usa a property
        filepath = os.path.join(settings.MEDIA_ROOT, 'processos', 'gerados', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        return filepath

    except Exception as e:
        print(f"Erro ao preencher mapa: {e}")
        raise


def gerar_planilha_odt(processo, dados_ai):
    """Gera arquivo ODT"""
    try:
        from odf.opendocument import OpenDocumentText
        from odf.text import P, H

        doc = OpenDocumentText()

        # Título
        doc.text.addElement(H(outlinelevel=1, text=f"Mapa Comparativo - Processo {processo.numero}"))

        # Informações
        doc.text.addElement(H(outlinelevel=2, text="Informações do Processo"))
        doc.text.addElement(P(text=f"Número: {processo.numero}"))
        doc.text.addElement(P(text=f"Descrição: {processo.descricao}"))
        doc.text.addElement(P(text=f"Valor Estimado: R$ {processo.valor_estimado:.2f}"))
        doc.text.addElement(P(text=f"Data: {processo.data_abertura.strftime('%d/%m/%Y')}"))
        doc.text.addElement(P(text=""))

        # Empresas
        empresas = dados_ai.get('empresas', [])
        if empresas:
            doc.text.addElement(H(outlinelevel=2, text="Empresas Participantes"))
            for emp in empresas:
                doc.text.addElement(P(text=f"- {emp.get('nome', 'N/A')}"))
                if emp.get('cnpj'):
                    doc.text.addElement(P(text=f"  CNPJ: {emp.get('cnpj')}"))
                # >>> NOVO: registra declínio/dúvida no relatório
                tipo = emp.get('tipo_resposta') or 'cotacao'
                if tipo != 'cotacao':
                    rotulo = 'Declinou' if tipo == 'declinio' else 'Apenas dúvida/esclarecimento'
                    motivo = emp.get('motivo_declinio') or ''
                    doc.text.addElement(P(text=f"  {rotulo}{': ' + motivo if motivo else ''}"))
                if emp.get('valor_global'):
                    doc.text.addElement(P(text=f"  Valor: {emp.get('valor_global')}"))
                doc.text.addElement(P(text=""))

        # Itens
        itens = dados_ai.get('itens', [])
        if itens:
            doc.text.addElement(H(outlinelevel=2, text="Itens do Processo"))
            for item in itens:
                doc.text.addElement(P(text=f"Item {item.get('item', '')}: {item.get('nome_em_portugues', '')}"))
                doc.text.addElement(P(text=f"  PI: {item.get('pi', '')}"))   # >>> NOVO
                doc.text.addElement(P(text=f"  Quantidade: {item.get('qtde', '')} {item.get('uf', '')}"))
                if item.get('valor_unitario_estimado'):
                    doc.text.addElement(P(text=f"  Valor Unitário: {item.get('valor_unitario_estimado')}"))
                if item.get('valor_total'):
                    doc.text.addElement(P(text=f"  Valor Total: {item.get('valor_total')}"))
                # Empresas do item
                empresas_data = item.get('empresas', {})
                if empresas_data:
                    doc.text.addElement(P(text="  Cotações:"))
                    for key, value in empresas_data.items():
                        if value:
                            doc.text.addElement(P(text=f"    {key}: {value}"))
                else:
                    doc.text.addElement(P(text="  Sem cotação"))   # >>> NOVO
                doc.text.addElement(P(text=""))

        # >>> NOVO: de onde veio cada dado — rota de extração, OCR local e custo
        extracao = dados_ai.get('extracao') or {}
        if extracao.get('arquivos'):
            doc.text.addElement(H(outlinelevel=2, text="Origem da Extração"))
            for arquivo in extracao['arquivos']:
                linha = f"- {arquivo.get('arquivo', '')}: {arquivo.get('rota') or 'n/d'}"
                if arquivo.get('status') and arquivo['status'] != 'success':
                    linha += f" [{arquivo['status']}: {arquivo.get('erro', '')}]"
                if arquivo.get('truncado'):
                    linha += " [resposta truncada por limite de tokens]"
                doc.text.addElement(P(text=linha))

            if extracao.get('com_ocr_local'):
                doc.text.addElement(P(text=""))
                doc.text.addElement(P(text=(
                    "ATENÇÃO: OCR local (Tesseract) foi aplicado em "
                    + ", ".join(extracao['com_ocr_local'])
                    + ". Preços e PI desses arquivos exigem conferência dígito a dígito "
                      "antes da homologação da pesquisa.")))

            doc.text.addElement(P(text=(
                f"Tokens: {extracao.get('tokens_entrada', 0)} entrada / "
                f"{extracao.get('tokens_saida', 0)} saída. "
                f"Custo estimado: USD {extracao.get('custo_usd', 0)}.")))
            doc.text.addElement(P(text=""))

        # >>> NOVO: perguntas e avisos ficam no documento, não só no JSON
        perguntas = dados_ai.get('perguntas') or []
        if perguntas:
            doc.text.addElement(H(outlinelevel=2, text="Perguntas dos Fornecedores"))
            for pergunta in perguntas:
                doc.text.addElement(P(text=f"- {pergunta.get('empresa', '')}: "
                                           f"{pergunta.get('pergunta', '')}"))
            doc.text.addElement(P(text=""))

        avisos = dados_ai.get('avisos_gerais') or []
        if avisos:
            doc.text.addElement(H(outlinelevel=2, text="Avisos do Processamento"))
            for aviso in avisos:
                doc.text.addElement(P(text=f"- {aviso}"))

        # Salva
        filename = f"mapa_comparativo_{processo.numero_slug}.odt"   # >>> ALTERADO: usa a property
        filepath = os.path.join(settings.MEDIA_ROOT, 'processos', 'gerados', filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        doc.save(filepath)
        return filepath

    except Exception as e:
        print(f"Erro ao gerar ODT: {e}")
        raise


# ==================== VIEWS DE CRIAÇÃO E DOWNLOAD ====================

@login_required   # >>> NOVO: a view usa request.user; sem isso, anônimo estoura
def novo_processo(request):
    """
    Duas etapas, no mesmo formulário. O NÚMERO do processo liga uma na outra.

    Etapa 1 - campo "modelo": o Modelo de Proposta (.xls/.xlsx) vira a LINHA DE
              BASE (itens e PI). Lido direto da planilha, sem IA.
    Etapa 2 - campo "file": pacote de respostas (.tgz/.zip) ou documento avulso.
              As cotações são casadas com a linha de base PELO PI.

    Os dois campos são opcionais, mas ao menos um tem de vir. Enviando só o
    modelo, o processo fica 'pendente' esperando as respostas. Enviando os dois
    de uma vez, as duas etapas rodam na mesma requisição.
    """
    if request.method == 'POST':
        numero = request.POST.get("numero")
        descricao = request.POST.get("descricao")
        valor_estimado = request.POST.get("valor_estimado")
        data_abertura = request.POST.get("data_abertura")
        modelo = request.FILES.get("modelo")     # >>> NOVO: linha de base
        arquivo = request.FILES.get("file")      # >>> ALTERADO: agora é opcional

        # >>> NOVO: devolve o que foi digitado quando a tela volta com erro
        formulario = {
            "form_numero": numero,
            "form_descricao": descricao,
            "form_valor_estimado": valor_estimado,
            "form_data_abertura": data_abertura,
        }

        def erro(mensagens):
            return render(request, "novoprocesso.html",
                          dict(formulario, erros=mensagens))

        erros = []

        # Validações
        if not numero or not descricao or not valor_estimado or not data_abertura:
            erros.append("Todos os campos são obrigatórios!")

        # >>> ALTERADO: sem o "if numero", campo vazio estourava TypeError no re.match
        if numero and not re.match(r'^[0-9./-]+$', numero):
            erros.append("Número deve conter apenas números, /, . e -")

        try:
            valor_estimado = float(valor_estimado)
            if valor_estimado < 0:
                erros.append("Valor estimado deve ser maior que zero.")
        except (TypeError, ValueError):   # >>> ALTERADO: float(None) levanta TypeError, não ValueError
            erros.append("Valor estimado deve ser um número.")

        # >>> ALTERADO: pelo menos um dos dois envios
        if not modelo and not arquivo:
            erros.append("Envie o Modelo de Proposta, o pacote de respostas, ou os dois.")

        for enviado, rotulo in ((modelo, "Modelo de Proposta"), (arquivo, "Arquivo de respostas")):
            if enviado and enviado.size > 50 * 1024 * 1024:
                erros.append(f"{rotulo}: arquivo muito grande. Máximo 50MB.")

        if modelo and os.path.splitext(modelo.name)[1].lower() not in ('.xls', '.xlsx', '.xlsm'):
            erros.append("O Modelo de Proposta precisa ser uma planilha .xls ou .xlsx.")

        if erros:
            return erro(erros)

        processo = None       # usados no except
        criado_agora = False
        try:
            # >>> ALTERADO: número repetido virou caminho normal (etapa 2)
            processo = Processo.objects.filter(numero=numero).first()

            if processo is not None:
                if processo.usuario_id not in (None, request.user.id):
                    return erro(["Esse número pertence a um processo de outro responsável."])
                if modelo and processo.itens.exists() and processo.fornecedores.exists():
                    return erro([
                        "Esse processo já tem linha de base e respostas processadas. "
                        "Para trocar o Modelo de Proposta, exclua o processo antes."])
                if not modelo and not processo.itens.exists():
                    return erro([
                        "Esse processo ainda não tem linha de base. Envie primeiro o "
                        "Modelo de Proposta com a coluna NÚMERO DE ESTOQUE (PI)."])
                processo.status = 'processando'
            else:
                if not modelo:
                    return erro([
                        "Processo novo começa pelo Modelo de Proposta: é ele que define "
                        "os itens e os PI da linha de base."])
                processo = Processo.objects.create(
                    usuario=request.user,
                    numero=numero,
                    descricao=descricao,
                    valor_estimado=valor_estimado,
                    data_abertura=data_abertura,
                    status='processando'
                )
                criado_agora = True

            # o pacote de respostas é o anexo principal; sem ele, guarda o modelo
            processo.arquivo_processo = arquivo or modelo
            processo.save()

            # Processa com IA
            contexto = {
                'numero': numero,
                'descricao': descricao,
                'valor_estimado': str(valor_estimado)
            }

            ai_processor = AIProcessor()
            emails_recebidos = 0   # >>> NOVO

            # Processa o arquivo
            with tempfile.TemporaryDirectory() as tmpdir:

                def gravar(upload):   # >>> NOVO: são dois uploads possíveis
                    caminho = os.path.join(tmpdir, upload.name)
                    with open(caminho, 'wb+') as destino:
                        for chunk in upload.chunks():
                            destino.write(chunk)
                    return caminho

                # ---------- ETAPA 1: linha de base ----------  >>> NOVO
                if modelo:
                    base = parse_modelo_proposta(gravar(modelo))
                    if not base or not any(item.get('pi') for item in base['itens']):
                        if criado_agora:
                            processo.delete()   # não deixa processo órfão travando o número
                            processo = None
                        return erro([
                            "Não consegui ler o Modelo de Proposta. Confira se é a planilha "
                            "enviada às empresas (com o cabeçalho ITEM / NÚMERO DE ESTOQUE / "
                            "NOMENCLATURA) e se a coluna NÚMERO DE ESTOQUE (PI) está preenchida."])
                    salvar_dados_ai(processo, base, modo='base')

                # sem respostas ainda: para aqui, aguardando as cotações
                if not arquivo:
                    processo.status = 'pendente'
                    processo.save(update_fields=['status'])
                    return redirect("processos")

                # ---------- ETAPA 2: respostas ----------  >>> NOVO
                base = linha_base(processo)
                file_path = gravar(arquivo)

                # Verifica se é compactado
                if arquivo.name.lower().endswith(('.zip', '.tgz', '.tar.gz', '.tar')):
                    extract_dir = os.path.join(tmpdir, 'extracted')
                    os.makedirs(extract_dir, exist_ok=True)

                    if file_path.endswith('.zip'):
                        with zipfile.ZipFile(file_path, 'r') as zip_ref:
                            zip_ref.extractall(extract_dir)
                    else:
                        import tarfile
                        mode = 'r:gz' if file_path.endswith(('.tgz', '.tar.gz')) else 'r'
                        with tarfile.open(file_path, mode) as tar_ref:
                            tar_ref.extractall(extract_dir)

                    emails_recebidos = contar_emails(extract_dir)   # >>> NOVO
                    results = ai_processor.process_directory(extract_dir, contexto, base)
                    dados_ai = ai_processor.merge_results(results, base)
                else:
                    # >>> ALTERADO: process_file (planilha não gasta IA) + merge com a base
                    resultado = ai_processor.process_file(file_path, contexto, base)
                    if arquivo.name.lower().endswith(('.eml', '.msg')):
                        emails_recebidos = 1
                    results = [{'filename': arquivo.name, 'ai_result': resultado}]
                    dados_ai = ai_processor.merge_results(results, base)

            # >>> NOVO: rota de extração de cada arquivo + custo, e os avisos que
            # dependem dela (OCR local, falha de leitura, resposta truncada)
            extracao = resumo_extracao(results)
            dados_ai['extracao'] = extracao
            avisos = dados_ai.setdefault('avisos_gerais', [])

            if extracao['com_ocr_local']:
                avisos.append(
                    "OCR local (Tesseract) aplicado em " +
                    ", ".join(extracao['com_ocr_local']) +
                    ": confira preços e PI desses arquivos dígito a dígito — "
                    "erro de OCR em número não é detectável pelo sistema.")
            for nome in extracao['com_falha']:
                detalhe = next((a['erro'] for a in extracao['arquivos']
                                if a['arquivo'] == nome and a['erro']), 'motivo não informado')
                avisos.append(f"{nome}: não foi possível extrair ({detalhe}).")
            for nome in extracao['truncados']:
                avisos.append(f"{nome}: resposta truncada por limite de tokens — "
                              f"itens podem ter ficado de fora.")

            # Salva dados da IA (JSON bruto, para auditoria)
            json_path = os.path.join(settings.MEDIA_ROOT, 'processos', 'gerados',
                                     f'dados_ai_{processo.numero_slug}.json')
            os.makedirs(os.path.dirname(json_path), exist_ok=True)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(dados_ai, f, ensure_ascii=False, indent=2)

            # >>> NOVO: grava os mesmos dados no SQLite (Fornecedor / Item / Cotacao)
            # modo='completo' preserva os itens da linha de base e casa pelo PI
            salvar_dados_ai(processo, dados_ai, emails_recebidos, modo='completo')

            # Gera arquivos
            xlsx_path = preencher_mapa_comparativo(processo, dados_ai)
            if xlsx_path:
                processo.arquivo_gerado_xlsx = f'processos/gerados/{os.path.basename(xlsx_path)}'

            odt_path = gerar_planilha_odt(processo, dados_ai)
            if odt_path:
                processo.arquivo_gerado_odt = f'processos/gerados/{os.path.basename(odt_path)}'

            processo.status = 'concluido'
            processo.save()

            return redirect("processos")

        except IntegrityError:
            return erro(["Não foi possível gravar o processo (número duplicado ou dado "
                         "inconsistente). Confira o número e tente de novo."])

        except Exception as e:
            # >>> NOVO: não deixa o processo travado em 'processando'
            if processo is not None and processo.pk:
                processo.status = 'erro'
                processo.save(update_fields=['status'])
            return erro([f"Erro ao processar processo: {str(e)}"])

    return render(request, "novoprocesso.html")


# >>> SUBSTITUÍDA: era "return render(request, 'documentos.html')"
def documentos(request):
    """Lista os processos com os dados consolidados pela IA."""
    numero = (request.GET.get("numero") or "").strip()
    status = (request.GET.get("status") or "").strip()

    consulta = Processo.objects.prefetch_related('fornecedores', 'itens__cotacoes')

    if numero:
        consulta = consulta.filter(numero__icontains=numero)
    if status and status != "todos":
        consulta = consulta.filter(status__iexact=status)

    paginator = Paginator(consulta, 5)
    pagina = paginator.get_page(request.GET.get("page"))

    # mantém os filtros ao trocar de página
    filtros = request.GET.copy()
    filtros.pop("page", None)

    return render(
        request,
        "documentos.html",
        {
            "resumos": [montar_resumo(processo) for processo in pagina],
            "pagina": pagina,
            "filtro_numero": numero,
            "filtro_status": status or "todos",
            "status_choices": Processo.STATUS_CHOICES,
            "querystring": filtros.urlencode(),
            "total_encontrados": paginator.count,
        }
    )


def mapas_gerados(request):
    return render(request, "mapasgerados.html")


# >>> ALTERADA: aceita também 'original' (o pacote de e-mails enviado)
def download_arquivo(request, tipo, processo_id):
    try:
        processo = Processo.objects.get(id=processo_id)
    except Processo.DoesNotExist:
        return HttpResponse("Processo não encontrado", status=404)

    caminhos = {
        'xlsx': os.path.join(settings.MEDIA_ROOT, str(processo.arquivo_gerado_xlsx)) if processo.arquivo_gerado_xlsx else None,
        'odt': os.path.join(settings.MEDIA_ROOT, str(processo.arquivo_gerado_odt)) if processo.arquivo_gerado_odt else None,
        'json': os.path.join(settings.MEDIA_ROOT, 'processos', 'gerados', f'dados_ai_{processo.numero_slug}.json'),
        'original': os.path.join(settings.MEDIA_ROOT, str(processo.arquivo_processo)) if processo.arquivo_processo else None,
    }

    caminho = caminhos.get(tipo)
    if caminho and os.path.exists(caminho):
        return FileResponse(open(caminho, 'rb'),
                            as_attachment=True,
                            filename=os.path.basename(caminho))

    return HttpResponse("Arquivo não disponível para este processo", status=404)
